"""
==========================================
Job Tracker Pro
Excel Handler
Developer : Yogendra Kumar
==========================================
"""

import os
from openpyxl import Workbook, load_workbook
from config import EXCEL_FILE


class ExcelHandler:

    def __init__(self):
        self.create_excel()

    # ----------------------------------------
    # Create Excel File
    # ----------------------------------------
    def create_excel(self):

        if not os.path.exists(EXCEL_FILE):

            workbook = Workbook()

            sheet = workbook.active
            sheet.title = "Jobs"

            headers = [
                "Company",
                "Position",
                "Location",
                "HR Name",
                "Email",
                "Phone",
                "Applied Date",
                "Status",
                "Source",
                "Resume",
                "Notes"
            ]

            sheet.append(headers)

            workbook.save(EXCEL_FILE)

    # ----------------------------------------
    # Add Job
    # ----------------------------------------
    def add_job(
        self,
        company,
        position,
        location,
        hr_name,
        email,
        phone,
        applied_date,
        status,
        source,
        resume,
        notes
    ):

        workbook = load_workbook(EXCEL_FILE)

        sheet = workbook["Jobs"]

        sheet.append([
            company,
            position,
            location,
            hr_name,
            email,
            phone,
            applied_date,
            status,
            source,
            resume,
            notes
        ])

        workbook.save(EXCEL_FILE)

    # ----------------------------------------
    # Get All Jobs
    # ----------------------------------------
    def get_all_jobs(self):

        workbook = load_workbook(EXCEL_FILE)

        sheet = workbook["Jobs"]

        jobs = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            jobs.append(row)

        return jobs