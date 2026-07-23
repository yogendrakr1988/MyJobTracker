    # screens/search_jobs.py

import os
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook

from config import EXCEL_FILE


class SearchJobsScreen:

    FILE = EXCEL_FILE

    def __init__(self, parent):
        self.parent = parent
        self.tree = None
        self.search_var = tk.StringVar()

    def search(self):

        for item in self.tree.get_children():
            self.tree.delete(item)

        if not os.path.exists(self.FILE):
            return

        wb = load_workbook(self.FILE)
        ws = wb.active

        keyword = self.search_var.get().strip().lower()

        rows = list(ws.iter_rows(values_only=True))

        for row in rows[1:]:

            text = " ".join(
                [str(x).lower() for x in row if x]
            )

            if keyword in text:

                self.tree.insert(
                    "",
                    "end",
                    values=(
                        row[0],
                        row[1],
                        row[2],
                        row[10],
                        row[8],
                        row[9]
                    )
                )

    def show(self):

        frame = ctk.CTkFrame(self.parent)
        frame.pack(fill="both", expand=True, padx=15, pady=15)

        ctk.CTkLabel(
            frame,
            text="Search Applications",
            font=("Arial",28,"bold")
        ).pack(pady=15)

        top = ctk.CTkFrame(frame)
        top.pack(fill="x", padx=20)

        ctk.CTkEntry(
            top,
            textvariable=self.search_var,
            width=350,
            placeholder_text="Company / Role / Status..."
        ).pack(side="left", padx=10)

        ctk.CTkButton(
            top,
            text="Search",
            command=self.search,
            width=140
        ).pack(side="left", padx=10)

        columns = (
            "Company",
            "Role",
            "Location",
            "Status",
            "Applied Date",
            "Follow-up"
        )

        self.tree = ttk.Treeview(
            frame,
            columns=columns,
            show="headings",
            height=18
        )

        for col in columns:

            self.tree.heading(col, text=col)
            self.tree.column(col, width=170, anchor="center")

        self.tree.pack(fill="both", expand=True, padx=20, pady=20)

        return frame