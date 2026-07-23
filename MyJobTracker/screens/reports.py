import os
from datetime import datetime
from collections import Counter

import customtkinter as ctk
from openpyxl import load_workbook, Workbook
from tkinter import messagebox

from config import REPORT_FOLDER, EXCEL_FILE


class ReportsScreen:

    FILE = EXCEL_FILE

    def __init__(self, parent):
        self.parent = parent

    def read_stats(self):

        result = {
            "Total": 0,
            "Applied": 0,
            "Interview": 0,
            "Selected": 0,
            "Rejected": 0
        }

        if not os.path.exists(self.FILE):
            return result

        wb = load_workbook(self.FILE)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        if len(rows) <= 1:
            return result

        header = rows[0]
        data = rows[1:]

        status_index = header.index("Status")

        counter = Counter(str(r[status_index]).strip() for r in data)

        result["Total"] = len(data)
        result["Applied"] = counter.get("Applied", 0)
        result["Interview"] = counter.get("Interview", 0)
        result["Selected"] = counter.get("Selected", 0)
        result["Rejected"] = counter.get("Rejected", 0)

        return result

    def export_report(self):

        if not os.path.exists(self.FILE):
            messagebox.showwarning("No Data", "No job data found to export.")
            return

        wb_src = load_workbook(self.FILE)
        ws_src = wb_src.active

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        stats = self.read_stats()

        ws.append(["Job Tracker Pro - Summary Report"])
        ws.append(["Generated on", datetime.now().strftime("%d-%m-%Y %I:%M %p")])
        ws.append([])
        ws.append(["Total Applications", stats["Total"]])
        ws.append(["Applied", stats["Applied"]])
        ws.append(["Interview", stats["Interview"]])
        ws.append(["Selected", stats["Selected"]])
        ws.append(["Rejected", stats["Rejected"]])
        ws.append([])
        ws.append(["All Applications"])

        for row in ws_src.iter_rows(values_only=True):
            ws.append(list(row))

        os.makedirs(REPORT_FOLDER, exist_ok=True)

        filename = f"Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(REPORT_FOLDER, filename)

        wb.save(filepath)

        messagebox.showinfo("Success", f"Report exported successfully:\n{filepath}")

    def card(self, parent, title, value, color="#1a1a1a"):

        box = ctk.CTkFrame(parent, width=160, height=110, fg_color="#f5f5f5", corner_radius=12)
        box.pack(side="left", padx=10)
        box.pack_propagate(False)

        ctk.CTkLabel(
            box, text=title, font=("Segoe UI", 14, "bold"), text_color="#555555"
        ).pack(pady=(18, 5))

        ctk.CTkLabel(
            box, text=str(value), font=("Segoe UI", 28, "bold"), text_color=color
        ).pack()

    def show(self):

        frame = ctk.CTkFrame(self.parent, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=20, pady=20)

        ctk.CTkLabel(
            frame,
            text="📊 Reports",
            font=("Segoe UI", 26, "bold")
        ).pack(pady=(0, 15), anchor="w")

        card = ctk.CTkFrame(frame, fg_color="white", corner_radius=15)
        card.pack(fill="both", expand=True)

        stats = self.read_stats()

        cards_row = ctk.CTkFrame(card, fg_color="white")
        cards_row.pack(pady=35)

        self.card(cards_row, "Total", stats["Total"])
        self.card(cards_row, "Applied", stats["Applied"], "#0d6efd")
        self.card(cards_row, "Interview", stats["Interview"], "#fd7e14")
        self.card(cards_row, "Selected", stats["Selected"], "#28a745")
        self.card(cards_row, "Rejected", stats["Rejected"], "#dc3545")

        ctk.CTkButton(
            card,
            text="⬇️  Export Report (Excel)",
            command=self.export_report,
            width=240,
            height=42,
            font=("Segoe UI", 14, "bold")
        ).pack(pady=20)

        return frame
