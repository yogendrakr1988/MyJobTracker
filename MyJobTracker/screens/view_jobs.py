# screens/view_jobs.py

import os
import sys
import subprocess

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook

from config import EXCEL_FILE


class ViewJobsScreen:

    FILE = EXCEL_FILE

    FIELDS = [
        "Company",
        "Role",
        "Location",
        "HR Name",
        "HR Email",
        "HR Mobile",
        "Job Portal",
        "Job Link",
        "Applied Date",
        "Follow-up Date",
        "Status",
        "Resume",
        "Notes"
    ]

    def __init__(self, parent):
        self.parent = parent
        self.tree = None
        self.row_map = {}
        self.full_data = {}
        self.checked_items = set()
        self.select_all_var = None

    def load_data(self):

        for item in self.tree.get_children():
            self.tree.delete(item)

        self.row_map = {}
        self.full_data = {}
        self.checked_items = set()

        if self.select_all_var is not None:
            self.select_all_var.set(False)

        if not os.path.exists(self.FILE):
            return

        wb = load_workbook(self.FILE)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        if len(rows) <= 1:
            return

        for i, row in enumerate(rows[1:]):

            excel_row_number = i + 2

            item_id = self.tree.insert(
                "",
                "end",
                text="☐",
                values=(
                    row[0],
                    row[1],
                    row[2],
                    row[10],
                    row[8],
                    row[9]
                )
            )

            self.row_map[item_id] = excel_row_number
            self.full_data[item_id] = row

    def refresh(self):
        self.load_data()

    def on_tree_click(self, event):
        """
        Tree ke column "#0" (checkbox wala column) par click hone par
        us row ki ✓ state toggle karta hai. Baaki columns par click
        normal row-selection ki tarah hi behave karta hai (koi change
        nahi).
        """
        region = self.tree.identify("region", event.x, event.y)
        col = self.tree.identify_column(event.x)

        if region in ("tree", "cell") and col == "#0":
            item_id = self.tree.identify_row(event.y)
            if item_id:
                self.toggle_check(item_id)
                return "break"  # row-selection highlight ko trigger na hone dein

    def toggle_check(self, item_id):
        if item_id in self.checked_items:
            self.checked_items.discard(item_id)
            self.tree.item(item_id, text="☐")
        else:
            self.checked_items.add(item_id)
            self.tree.item(item_id, text="☑")

        self._sync_select_all_checkbox()

    def _sync_select_all_checkbox(self):
        if self.select_all_var is None:
            return
        all_items = self.tree.get_children()
        all_checked = bool(all_items) and all(i in self.checked_items for i in all_items)
        self.select_all_var.set(all_checked)

    def toggle_select_all(self):
        want_check = bool(self.select_all_var.get())
        for item_id in self.tree.get_children():
            if want_check:
                self.checked_items.add(item_id)
                self.tree.item(item_id, text="☑")
            else:
                self.checked_items.discard(item_id)
                self.tree.item(item_id, text="☐")

    def get_selected_job_rows(self):
        """Tick ki hui rows ko job_data dicts (FIELDS keys ke saath) mein convert karta hai."""
        rows = []
        for item_id in self.checked_items:
            raw = self.full_data.get(item_id)
            if not raw:
                continue
            rows.append(dict(zip(self.FIELDS, raw)))
        return rows

    def bulk_send_email(self):
        rows = self.get_selected_job_rows()

        if not rows:
            messagebox.showwarning(
                "No Selection",
                "Pehle list mein ✓ (tick) karke kam se kam ek application select karein."
            )
            return

        no_email = [r for r in rows if not (r.get("HR Email") or "").strip()]

        confirm_msg = f"{len(rows)} application(s) ko Email bheja jayega."
        if no_email:
            confirm_msg += f"\n({len(no_email)} mein HR Email khaali hai — woh skip ho jayengi.)"

        if not messagebox.askyesno("Confirm Bulk Email", confirm_msg + "\n\nContinue karein?"):
            return

        from screens.bulk_send import run_bulk_send
        run_bulk_send(self.parent, rows, mode="email")

    def bulk_send_whatsapp(self):
        rows = self.get_selected_job_rows()

        if not rows:
            messagebox.showwarning(
                "No Selection",
                "Pehle list mein ✓ (tick) karke kam se kam ek application select karein."
            )
            return

        confirm = messagebox.askyesno(
            "WhatsApp Bulk Send — Risk Warning",
            "⚠️ WhatsApp Web par lagatar bahut saare messages bhejne se aapka "
            "number temporarily/permanently block ho sakta hai — yeh WhatsApp ki "
            "apni policy hai, isse bachne ka koi 100% guaranteed tareeka nahi hai.\n\n"
            f"Aap {len(rows)} HR(s) ko WhatsApp bhejne ja rahe hain. Har message ke "
            "beech jaan-bujh kar delay rakha jayega taaki risk kam ho, lekin risk "
            "poori tarah khatam nahi hota.\n\n"
            "Agar aap sure nahi hain, to iske bajaay 'Bulk Email' use karna zyada safe hai.\n\n"
            "Kya aap WhatsApp bulk send ke saath continue karna chahte hain?"
        )

        if not confirm:
            return

        from screens.bulk_send import run_bulk_send
        run_bulk_send(self.parent, rows, mode="whatsapp")

    def delete_selected(self):

        selected = self.tree.selection()

        if not selected:
            messagebox.showwarning("No Selection", "Please select a job to delete.")
            return

        confirm = messagebox.askyesno(
            "Confirm Delete",
            "Are you sure you want to delete the selected application(s)?\nThis cannot be undone."
        )

        if not confirm:
            return

        if not os.path.exists(self.FILE):
            return

        wb = load_workbook(self.FILE)
        ws = wb.active

        row_numbers = sorted(
            (self.row_map[item] for item in selected),
            reverse=True
        )

        for row_number in row_numbers:
            ws.delete_rows(row_number)

        wb.save(self.FILE)

        messagebox.showinfo("Deleted", "Selected application(s) deleted successfully.")

        self.load_data()

    def open_resume(self):

        selected = self.tree.selection()

        if not selected:
            messagebox.showwarning("No Selection", "Please select a job first.")
            return

        row = self.full_data.get(selected[0])

        if not row:
            return

        resume_path = row[11]

        if not resume_path or not os.path.exists(str(resume_path)):
            messagebox.showinfo("No Resume", "No resume attached (or file not found) for this application.")
            return

        try:
            if sys.platform.startswith("win"):
                os.startfile(resume_path)
            elif sys.platform == "darwin":
                subprocess.call(["open", resume_path])
            else:
                subprocess.call(["xdg-open", resume_path])

        except Exception as e:
            messagebox.showerror("Error", f"Could not open resume:\n{e}")

    def load_in_form(self):

        selected = self.tree.selection()

        if not selected:
            messagebox.showwarning("No Selection", "Please select a job to load.")
            return

        row = self.full_data.get(selected[0])

        if not row:
            return

        data = dict(zip(self.FIELDS, row))

        if hasattr(self.parent, "show_add_job"):
            self.parent.show_add_job(prefill=data)
        else:
            messagebox.showerror("Error", "Could not open Add Job screen.")

    def edit_selected(self):

        selected = self.tree.selection()

        if not selected:
            messagebox.showwarning("No Selection", "Please select a job to edit.")
            return

        item_id = selected[0]
        row_number = self.row_map[item_id]
        row = self.full_data[item_id]

        win = ctk.CTkToplevel(self.parent)
        win.title("Edit Application")
        win.geometry("520x600")
        win.grab_set()

        scroll = ctk.CTkScrollableFrame(win, fg_color="white")
        scroll.pack(fill="both", expand=True, padx=15, pady=15)

        entries = {}

        for i, field in enumerate(self.FIELDS):

            ctk.CTkLabel(
                scroll,
                text=field,
                font=("Segoe UI", 13, "bold"),
                text_color="#1a1a1a"
            ).grid(row=i, column=0, sticky="w", pady=8, padx=(0, 10))

            value = row[i] if i < len(row) and row[i] is not None else ""

            var = tk.StringVar(value=str(value))
            entries[field] = var

            ctk.CTkEntry(
                scroll,
                textvariable=var,
                width=280,
                font=("Segoe UI", 13)
            ).grid(row=i, column=1, pady=8)

        def save_edit():

            wb = load_workbook(self.FILE)
            ws = wb.active

            for col_index, field in enumerate(self.FIELDS, start=1):
                ws.cell(row=row_number, column=col_index, value=entries[field].get().strip())

            wb.save(self.FILE)

            win.destroy()

            self.load_data()

            messagebox.showinfo("Success", "Application updated successfully.")

        ctk.CTkButton(
            win,
            text="💾 Save Changes",
            command=save_edit,
            width=200,
            height=40,
            fg_color="#28a745",
            hover_color="#218838",
            font=("Segoe UI", 14, "bold")
        ).pack(pady=15)

    def show(self):

        frame = ctk.CTkFrame(self.parent)
        frame.pack(fill="both", expand=True, padx=15, pady=15)

        ctk.CTkLabel(
            frame,
            text="All Applications",
            font=("Arial", 28, "bold")
        ).pack(pady=15)

        table_frame = ctk.CTkFrame(frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)

        select_all_row = ctk.CTkFrame(table_frame, fg_color="transparent")
        select_all_row.pack(fill="x", pady=(0, 6))

        self.select_all_var = tk.BooleanVar(value=False)

        ctk.CTkCheckBox(
            select_all_row,
            text="Select All (tick karke bulk Email / WhatsApp bhejein)",
            variable=self.select_all_var,
            command=self.toggle_select_all,
            font=("Segoe UI", 13)
        ).pack(side="left")

        columns = (
            "Company",
            "Role",
            "Location",
            "Status",
            "Applied Date",
            "Follow-up"
        )

        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="tree headings",
            height=18
        )

        self.tree.heading("#0", text="✓")
        self.tree.column("#0", width=40, anchor="center", stretch=False)

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=170, anchor="center")

        self.tree.bind("<Button-1>", self.on_tree_click)

        scrollbar = ttk.Scrollbar(
            table_frame,
            orient="vertical",
            command=self.tree.yview
        )

        self.tree.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

        button_frame = ctk.CTkFrame(frame)
        button_frame.pack(pady=15)

        ctk.CTkButton(
            button_frame,
            text="Refresh",
            width=140,
            command=self.refresh
        ).pack(side="left", padx=8)

        ctk.CTkButton(
            button_frame,
            text="Delete",
            width=140,
            fg_color="#dc3545",
            hover_color="#c82333",
            command=self.delete_selected
        ).pack(side="left", padx=8)

        ctk.CTkButton(
            button_frame,
            text="Edit",
            width=140,
            command=self.edit_selected
        ).pack(side="left", padx=8)

        ctk.CTkButton(
            button_frame,
            text="Open Resume",
            width=140,
            command=self.open_resume
        ).pack(side="left", padx=8)

        ctk.CTkButton(
            button_frame,
            text="Load in Add Job",
            width=150,
            fg_color="#6f42c1",
            hover_color="#5a349b",
            command=self.load_in_form
        ).pack(side="left", padx=8)

        ctk.CTkButton(
            button_frame,
            text="📧 Bulk Email (Selected)",
            width=190,
            fg_color="#0d6efd",
            hover_color="#0b5ed7",
            command=self.bulk_send_email
        ).pack(side="left", padx=8)

        ctk.CTkButton(
            button_frame,
            text="💬 Bulk WhatsApp (Selected)",
            width=200,
            fg_color="#198754",
            hover_color="#146c43",
            command=self.bulk_send_whatsapp
        ).pack(side="left", padx=8)

        self.load_data()

        return frame
