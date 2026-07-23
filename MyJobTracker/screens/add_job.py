import os
import threading

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook

from notifier import send_email, send_whatsapp, send_developer_copy
from settings_manager import load_settings
from config import EXCEL_FILE as FILE
RESUME_FOLDER = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Resume")
os.makedirs(RESUME_FOLDER, exist_ok=True)

LABEL_FONT = ("Segoe UI", 15, "bold")
ENTRY_FONT = ("Segoe UI", 14)
HEADING_FONT = ("Segoe UI", 26, "bold")


def _show_success_popup(title, msg, parent, auto_close_ms=2200):
    """
    Success popup jo khud-ba-khud band ho jaata hai (auto_close_ms ke
    baad) — koi manual OK click ki zaroorat nahi. Agar user chahe toh
    OK button se turant bhi band kar sakta hai.

    (Failures/warnings ke liye jaan-boojh kar messagebox.showwarning hi
    use kiya jaata hai — woh manually band karna padta hai, taaki koi
    error accidentally miss na ho.)
    """

    popup = ctk.CTkToplevel(parent)
    popup.title(title)
    popup.resizable(False, False)
    popup.transient(parent)
    popup.grab_set()
    popup.attributes("-topmost", True)

    ctk.CTkLabel(
        popup,
        text=f"✅  {msg}",
        font=("Segoe UI", 13),
        wraplength=360,
        justify="left"
    ).pack(padx=24, pady=(22, 14))

    ctk.CTkButton(
        popup,
        text="OK",
        width=90,
        command=popup.destroy
    ).pack(pady=(0, 18))

    # Center the popup roughly over its parent window.
    #
    # IMPORTANT: a single update_idletasks() right after packing is not
    # enough - CTk widgets (canvas based, with wraplength text) often
    # haven't finished computing their real layout yet on the very
    # first pass, so winfo_width()/winfo_height() report a too-small
    # size and the popup gets geometry-locked at that tiny size. Doing
    # a couple of update passes first, and reading the *requested*
    # size (reqwidth/reqheight) instead of the current size, gives the
    # real, fully-laid-out dimensions.
    popup.update_idletasks()
    popup.update()

    w = max(popup.winfo_reqwidth(), popup.winfo_width())
    h = max(popup.winfo_reqheight(), popup.winfo_height())

    px = parent.winfo_rootx() + (parent.winfo_width() - w) // 2
    py = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
    popup.geometry(f"{w}x{h}+{max(px, 0)}+{max(py, 0)}")

    # Belt-and-braces: one more re-check a moment later, in case the
    # very first frame still rendered small on a slower machine.
    def _fix_size_once_more():
        try:
            if not popup.winfo_exists():
                return
            w2 = max(popup.winfo_reqwidth(), popup.winfo_width())
            h2 = max(popup.winfo_reqheight(), popup.winfo_height())
            if w2 > w or h2 > h:
                px2 = parent.winfo_rootx() + (parent.winfo_width() - w2) // 2
                py2 = parent.winfo_rooty() + (parent.winfo_height() - h2) // 2
                popup.geometry(f"{w2}x{h2}+{max(px2, 0)}+{max(py2, 0)}")
        except Exception:
            pass

    popup.after(60, _fix_size_once_more)

    def _auto_close():
        try:
            if popup.winfo_exists():
                popup.destroy()
        except Exception:
            pass

    popup.after(auto_close_ms, _auto_close)


class AddJobScreen:

    def __init__(self, parent):
        self.parent = parent
        self.entries = {}
        self.status = tk.StringVar(value="Applied")
        self.resume_var = tk.StringVar()
        self.status_label = None
        self.status_lines = {"email": "", "whatsapp": ""}

        # Guard flag so the Save button (or Ctrl+S) can never fire two
        # saves for the same click / keypress. Only one save can be
        # "in flight" at a time.
        self._is_saving = False
        self.save_btn = None

    def show(self):

        frame = ctk.CTkFrame(self.parent, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=20, pady=15)

        ctk.CTkLabel(
            frame,
            text="➕  Add New Job Application",
            font=HEADING_FONT
        ).pack(pady=(0, 15), anchor="w")

        # White card holds the whole form - no dark gap in the middle
        form_card = ctk.CTkFrame(frame, fg_color="white", corner_radius=15)
        form_card.pack(fill="both", expand=True)

        left = ctk.CTkFrame(form_card, fg_color="white")
        right = ctk.CTkFrame(form_card, fg_color="white")

        left.pack(side="left", fill="both", expand=True, padx=30, pady=25)
        right.pack(side="right", fill="both", expand=True, padx=30, pady=25)

        left_fields = [
            "Company",
            "Role",
            "Location",
            "HR Name",
            "HR Email",
            "HR Mobile"
        ]

        right_fields = [
            "Applied Date",
            "Follow-up Date",
            "Job Portal",
            "Job Link"
        ]

        for i, field in enumerate(left_fields):

            ctk.CTkLabel(
                left,
                text=field,
                font=LABEL_FONT,
                text_color="#1a1a1a",
                anchor="w"
            ).grid(row=i, column=0, sticky="w", pady=10, padx=(0, 12))

            var = tk.StringVar()

            self.entries[field] = var

            e = ctk.CTkEntry(
                left,
                textvariable=var,
                width=280,
                font=ENTRY_FONT
            )
            e.grid(row=i, column=1, pady=10)

            if field == "Company":
                self.company_entry = e

        ctk.CTkLabel(
            right,
            text="Status",
            font=LABEL_FONT,
            text_color="#1a1a1a",
            anchor="w"
        ).grid(row=0, column=0, sticky="w", pady=10, padx=(0, 12))

        ctk.CTkComboBox(
            right,
            variable=self.status,
            values=["Applied", "Interview", "Selected", "Rejected"],
            width=280,
            font=ENTRY_FONT,
            dropdown_font=ENTRY_FONT
        ).grid(row=0, column=1, pady=10)

        self.entries["Status"] = self.status

        row = 1

        for field in right_fields:

            ctk.CTkLabel(
                right,
                text=field,
                font=LABEL_FONT,
                text_color="#1a1a1a",
                anchor="w"
            ).grid(row=row, column=0, sticky="w", pady=10, padx=(0, 12))

            var = tk.StringVar()

            self.entries[field] = var

            ctk.CTkEntry(
                right,
                textvariable=var,
                width=280,
                font=ENTRY_FONT
            ).grid(row=row, column=1, pady=10)

            row += 1

        ctk.CTkLabel(
            right,
            text="Resume",
            font=LABEL_FONT,
            text_color="#1a1a1a",
            anchor="w"
        ).grid(row=row, column=0, sticky="w", pady=10, padx=(0, 12))

        resume_row = ctk.CTkFrame(right, fg_color="white")
        resume_row.grid(row=row, column=1, sticky="w", pady=10)

        ctk.CTkEntry(
            resume_row,
            textvariable=self.resume_var,
            width=195,
            font=ENTRY_FONT
        ).pack(side="left")

        ctk.CTkButton(
            resume_row,
            text="Attach",
            command=self.attach,
            width=75,
            font=ENTRY_FONT
        ).pack(side="left", padx=(8, 0))

        row += 1

        ctk.CTkLabel(
            right,
            text="Notes",
            font=LABEL_FONT,
            text_color="#1a1a1a",
            anchor="nw"
        ).grid(row=row, column=0, sticky="nw", pady=10, padx=(0, 12))

        self.notes = ctk.CTkTextbox(
            right,
            width=280,
            height=90,
            font=ENTRY_FONT
        )

        self.notes.grid(row=row, column=1, pady=10)

        self.entries["Notes"] = self.notes

        # Status line under the form always mirrors the processing window,
        # so the current state is visible even if the popup is closed.
        self.status_label = ctk.CTkLabel(
            frame,
            text="",
            font=("Segoe UI", 12),
            text_color="#333333",
            justify="left"
        )
        self.status_label.pack(fill="x", padx=5, pady=(10, 0), anchor="w")

        btn = ctk.CTkFrame(frame, fg_color="transparent")
        btn.pack(side="bottom", pady=(8, 0))

        self.save_btn = ctk.CTkButton(
            btn,
            text="💾 Save",
            command=self.save,
            width=150,
            height=36,
            fg_color="#28a745",
            hover_color="#218838",
            font=("Segoe UI", 13, "bold")
        )
        self.save_btn.pack(side="left", padx=8)

        ctk.CTkButton(
            btn,
            text="❌ Exit",
            command=frame.destroy,
            width=150,
            height=36,
            fg_color="#dc3545",
            hover_color="#c82333",
            font=("Segoe UI", 13, "bold")
        ).pack(side="left", padx=8)

        self.company_entry.focus_set()

        return frame

    # ==========================================
    # Save (writes the Excel row ONCE, then kicks
    # off the visible email / WhatsApp processing)
    # ==========================================

    def save(self):

        # Guard: ignore a second click / Ctrl+S while a save is already
        # running, so the row is never written twice for one submission.
        if self._is_saving:
            return

        vals = {}

        for key, widget in self.entries.items():

            if isinstance(widget, ctk.CTkTextbox):
                vals[key] = widget.get("1.0", "end").strip()

            else:
                vals[key] = widget.get().strip()

        vals["Resume"] = self.resume_var.get().strip()

        if vals["Company"] == "":
            messagebox.showerror(
                "Validation Error",
                "Company Name is required."
            )
            return

        if vals["Role"] == "":
            messagebox.showerror(
                "Validation Error",
                "Role is required."
            )
            return

        self._is_saving = True

        if self.save_btn is not None:
            self.save_btn.configure(state="disabled", text="⏳ Saving...")

        try:

            if os.path.exists(FILE):

                wb = load_workbook(FILE)
                ws = wb.active

            else:

                wb = Workbook()
                ws = wb.active

                ws.append([
                    "Company",
                    "Role",
                    "Location",
                    "HR Name",
                    "HR Email",
                    "HR Mobile",
                    "Job Portal",
                    "Job Link",
                    "Applied Date",
                    "Follow-up Date",
                    "Status",
                    "Resume",
                    "Notes"
                ])

            ws.append([
                vals.get("Company", ""),
                vals.get("Role", ""),
                vals.get("Location", ""),
                vals.get("HR Name", ""),
                vals.get("HR Email", ""),
                vals.get("HR Mobile", ""),
                vals.get("Job Portal", ""),
                vals.get("Job Link", ""),
                vals.get("Applied Date", ""),
                vals.get("Follow-up Date", ""),
                vals.get("Status", ""),
                vals.get("Resume", ""),
                vals.get("Notes", "")
            ])

            wb.save(FILE)

        except Exception as e:

            self._is_saving = False

            if self.save_btn is not None:
                self.save_btn.configure(state="normal", text="💾 Save")

            messagebox.showerror(
                "Save Failed",
                f"Could not save the application to Excel:\n{e}"
            )
            return

        # Data is now safely on disk (written exactly once). Everything
        # after this point is just visible progress / notifications.
        self.open_processing_window(dict(vals))

    # ==========================================
    # Processing window: shows EVERYTHING live
    # (save confirmation, email progress, WhatsApp
    # progress) and pops a SEPARATE success/warning
    # message for email and for WhatsApp the moment
    # each one finishes - nothing runs hidden.
    # ==========================================

    def open_processing_window(self, vals):

        self.status_lines = {"email": "📤 Sending email...", "whatsapp": "📤 Sending WhatsApp..."}

        if self.status_label is not None and self.status_label.winfo_exists():
            self.status_label.configure(text=self._status_text())

        win = ctk.CTkToplevel(self.parent)
        win.title("Application Progress")
        win.geometry("500x420")
        win.minsize(480, 380)
        win.resizable(False, True)  # can't widen, but CAN be made taller if content overflows
        win.transient(self.parent)
        win.grab_set()
        win.protocol("WM_DELETE_WINDOW", lambda: None)  # closes only via Close button once done

        ctk.CTkLabel(
            win,
            text="📨 Processing Your Application",
            font=("Segoe UI", 18, "bold")
        ).pack(pady=(22, 4))

        ctk.CTkLabel(
            win,
            text=f'{vals.get("Company", "")} — {vals.get("Role", "")}',
            font=("Segoe UI", 13),
            text_color="#666666"
        ).pack(pady=(0, 18))

        saved_row = ctk.CTkLabel(
            win,
            text="✅ Data saved to jobs.xlsx",
            font=("Segoe UI", 14, "bold"),
            text_color="#28a745",
            anchor="w"
        )
        saved_row.pack(fill="x", padx=30, pady=6)

        email_row = ctk.CTkLabel(
            win,
            text="📧 Email: ⏳ Sending...",
            font=("Segoe UI", 14),
            anchor="w",
            justify="left",
            wraplength=420
        )
        email_row.pack(fill="x", padx=30, pady=6)

        whatsapp_row = ctk.CTkLabel(
            win,
            text="💬 WhatsApp: ⏳ Sending...",
            font=("Segoe UI", 14),
            anchor="w",
            justify="left",
            wraplength=420
        )
        whatsapp_row.pack(fill="x", padx=30, pady=6)

        # Developer Copy row is ALWAYS visible so it's never hidden - if the
        # user hasn't opted in, it just says so plainly and nothing is sent.
        dev_consent = bool(load_settings().get("developer_copy_consent", False))

        dev_row = ctk.CTkLabel(
            win,
            text=("📨 Developer Copy: ⏳ Sending..." if dev_consent
                  else "📨 Developer Copy: ⏸ Off (enable in Settings if you want this)"),
            font=("Segoe UI", 14),
            anchor="w",
            justify="left",
            wraplength=420
        )
        dev_row.pack(fill="x", padx=30, pady=6)

        close_btn = ctk.CTkButton(
            win,
            text="Please wait...",
            state="disabled",
            width=140,
            height=36,
            font=("Segoe UI", 13, "bold")
        )
        close_btn.pack(pady=(22, 18))

        # If Developer Copy is off, there is nothing to wait for on that
        # channel - it's already "done" and stays done.
        self._notif_done = {"email": False, "whatsapp": False, "developer": not dev_consent}
        # Tracks whether each channel actually SUCCEEDED (not just
        # finished) - used to decide if the whole window can safely
        # auto-close, or if it must stay open for the user to notice a
        # failure.
        self._notif_ok = {"email": True, "whatsapp": True, "developer": True}

        def finish_saving_state():
            # Once ALL channels are done, re-enable Save / clear the
            # form for the next entry.
            self._is_saving = False

            if self.save_btn is not None and self.save_btn.winfo_exists():
                self.save_btn.configure(state="normal", text="💾 Save")

            self.clear()

            if hasattr(self, "company_entry") and self.company_entry.winfo_exists():
                self.company_entry.focus_set()

            if not win.winfo_exists():
                return

            all_ok = (
                self._notif_ok["email"]
                and self._notif_ok["whatsapp"]
                and self._notif_ok["developer"]
            )

            # Safety net: re-enable the window's own X button (and
            # Alt+F4) now that everything is done, in case auto-close
            # below doesn't fire for any reason.
            win.protocol("WM_DELETE_WINDOW", win.destroy)

            if all_ok:
                # Sab kuch successful — is poori window ko bhi khud-ba-khud
                # band kar dete hain, koi manual click ki zaroorat nahi.
                close_btn.configure(state="normal", text="✅ All done — closing...")
                win.after(1800, lambda: win.destroy() if win.winfo_exists() else None)
            else:
                # Kam se kam ek channel fail hua — window manually band
                # karni hogi, taaki failure kabhi accidentally miss na ho.
                close_btn.configure(state="normal", text="Close", command=win.destroy)

        def maybe_finish():
            if self._notif_done["email"] and self._notif_done["whatsapp"] and self._notif_done["developer"]:
                finish_saving_state()

        def update_line(key, ok, msg):

            def apply():

                icon = "✅" if ok else "⚠️"
                self.status_lines[key] = f"{icon} {msg}"
                self._notif_done[key] = True
                self._notif_ok[key] = ok

                if self.status_label is not None and self.status_label.winfo_exists():
                    self.status_label.configure(text=self._status_text())

                if win.winfo_exists():

                    if key == "email":
                        email_row.configure(text=f"📧 Email: {icon} {msg}")
                    elif key == "whatsapp":
                        whatsapp_row.configure(text=f"💬 WhatsApp: {icon} {msg}")
                    else:
                        dev_row.configure(text=f"📨 Developer Copy: {icon} {msg}")

                    # FIX: agar status message lamba ho (jaise error ke
                    # saath screenshot path), toh window ki height ko
                    # content ke hisaab se badha dete hain — warna neeche
                    # wala "Close" button window ke bahar push ho kar
                    # chhup jaata tha. Width waisi hi fixed rehti hai,
                    # sirf height content-fit hoti hai.
                    win.update_idletasks()
                    needed_h = max(win.winfo_reqheight(), 380)
                    current_w = win.winfo_width() or 500
                    win.geometry(f"{current_w}x{needed_h}")

                    # Separate popup for THIS channel only, right when it
                    # finishes - email, WhatsApp, and developer copy never
                    # get merged into one combined message.
                    title = {
                        "email": "Email Status",
                        "whatsapp": "WhatsApp Status",
                        "developer": "Developer Copy Status"
                    }[key]

                    if ok:
                        _show_success_popup(title, msg, win)
                    else:
                        messagebox.showwarning(title, msg, parent=win)

                maybe_finish()

            self.parent.after(0, apply)

        def email_worker():
            try:
                ok, msg = send_email(vals)
            except BaseException as e:
                ok, msg = False, f"Email failed (unexpected): {e}"
            update_line("email", ok, msg)

        def whatsapp_worker():
            try:
                ok, msg = send_whatsapp(vals)
            except BaseException as e:
                ok, msg = False, f"WhatsApp failed (unexpected): {e}"
            update_line("whatsapp", ok, msg)

        def developer_worker():
            try:
                ok, msg = send_developer_copy(vals)
            except BaseException as e:
                ok, msg = False, f"Developer copy failed (unexpected): {e}"
            update_line("developer", ok, msg)

        threading.Thread(target=email_worker, daemon=True).start()
        threading.Thread(target=whatsapp_worker, daemon=True).start()

        # Only runs a background send if the user has actually opted in.
        if dev_consent:
            threading.Thread(target=developer_worker, daemon=True).start()

    def _status_text(self):

        return "\n".join([
            self.status_lines.get("email", ""),
            self.status_lines.get("whatsapp", "")
        ])

    def clear(self):

        for key, widget in self.entries.items():

            if isinstance(widget, ctk.CTkTextbox):

                widget.delete("1.0", "end")

            else:

                widget.set("")

        self.status.set("Applied")

        self.resume_var.set("")

    def attach(self):

        # Default folder: dialog ab hamesha app ke apne Resume/ folder se
        # khulta hai (agar wahan koi file pehle se maujood hai), taaki
        # galti se Downloads ya kisi aur temporary folder ki file select
        # na ho jaaye - jo baad mein move/delete/rename hone par resume
        # attach fail kar deti thi.
        current = self.resume_var.get().strip()
        if current and os.path.isdir(os.path.dirname(current)):
            initial_dir = os.path.dirname(current)
        else:
            initial_dir = RESUME_FOLDER

        file = filedialog.askopenfilename(
            initialdir=initial_dir,
            filetypes=[
                ("PDF Files", "*.pdf"),
                ("Word Files", "*.doc *.docx"),
                ("All Files", "*.*")
            ]
        )

        if file:

            self.resume_var.set(file)

    def refresh(self):

        self.clear()

        self.company_entry.focus_set()

    def load_defaults(self):

        from datetime import datetime

        today = datetime.now().strftime("%d-%m-%Y")

        self.entries["Applied Date"].set(today)

        self.entries["Follow-up Date"].set(today)

    def destroy(self):

        for widget in self.parent.winfo_children():

            widget.destroy()

    def hide(self):

        for widget in self.parent.winfo_children():

            widget.pack_forget()

    def create_excel(self):

        if os.path.exists(FILE):
            return

        wb = Workbook()
        ws = wb.active

        ws.title = "Job Applications"

        ws.append([
            "Company",
            "Role",
            "Location",
            "HR Name",
            "HR Email",
            "HR Mobile",
            "Job Portal",
            "Job Link",
            "Applied Date",
            "Follow-up Date",
            "Status",
            "Resume",
            "Notes"
        ])

        wb.save(FILE)

    def initialize(self):

        self.create_excel()

        self.load_defaults()

        self.company_entry.focus_set()

        return self

    def get_form_data(self):

        data = {}

        for key, widget in self.entries.items():

            if isinstance(widget, ctk.CTkTextbox):

                data[key] = widget.get("1.0", "end").strip()

            else:

                data[key] = widget.get().strip()

        data["Resume"] = self.resume_var.get().strip()

        return data

    def set_form_data(self, data):

        self.clear()

        for key, value in data.items():

            if key not in self.entries:
                continue

            widget = self.entries[key]

            if isinstance(widget, ctk.CTkTextbox):

                widget.insert("1.0", value)

            else:

                widget.set(value)

        if "Resume" in data:

            self.resume_var.set(data["Resume"])

    def validate(self):

        if self.entries["Company"].get().strip() == "":
            messagebox.showerror(
                "Validation Error",
                "Company Name is required."
            )
            self.company_entry.focus_set()
            return False

        if self.entries["Role"].get().strip() == "":
            messagebox.showerror(
                "Validation Error",
                "Role is required."
            )
            return False

        if self.entries["Applied Date"].get().strip() == "":
            messagebox.showerror(
                "Validation Error",
                "Applied Date is required."
            )
            return False

        return True

    def save_and_refresh(self):

        if not self.validate():
            return

        self.save()

    def export_to_dict(self):

        return {
            "Company": self.entries["Company"].get().strip(),
            "Role": self.entries["Role"].get().strip(),
            "Location": self.entries["Location"].get().strip(),
            "HR Name": self.entries["HR Name"].get().strip(),
            "HR Email": self.entries["HR Email"].get().strip(),
            "HR Mobile": self.entries["HR Mobile"].get().strip(),
            "Applied Date": self.entries["Applied Date"].get().strip(),
            "Follow-up Date": self.entries["Follow-up Date"].get().strip(),
            "Job Portal": self.entries["Job Portal"].get().strip(),
            "Job Link": self.entries["Job Link"].get().strip(),
            "Status": self.status.get(),
            "Resume": self.resume_var.get().strip(),
            "Notes": self.notes.get("1.0", "end").strip()
        }

    def reset_form(self):

        self.clear()

        self.load_defaults()

        self.company_entry.focus_set()

    def fill_form(self, values):

        self.clear()

        for key in self.entries:

            if key not in values:
                continue

            widget = self.entries[key]

            if isinstance(widget, ctk.CTkTextbox):

                widget.insert("1.0", values[key])

            else:

                widget.set(values[key])

        if "Resume" in values:

            self.resume_var.set(values["Resume"])

        self.company_entry.focus_set()

    def is_empty(self):

        for key, widget in self.entries.items():

            if isinstance(widget, ctk.CTkTextbox):

                if widget.get("1.0", "end").strip():
                    return False

            else:

                if widget.get().strip():
                    return False

        return True

    def close(self):

        if self.is_empty():

            return True

        result = messagebox.askyesno(
            "Confirm Exit",
            "Unsaved data found.\n\nDo you want to exit?"
        )

        return result

    def bind_shortcuts(self, root):

        root.bind("<Control-s>", lambda e: self.save())
        root.bind("<Control-n>", lambda e: self.reset_form())
        root.bind("<Escape>", lambda e: root.destroy())

    def __del__(self):

        self.entries.clear()

    def __repr__(self):

        return "AddJobScreen()"

    def __str__(self):

        return "Job Tracker - Add Job Screen"

    def __enter__(self):

        return self

    def __exit__(self, exc_type, exc_value, traceback):

        return False
