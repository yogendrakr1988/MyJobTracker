import os
from datetime import datetime
from collections import Counter

import customtkinter as ctk
from openpyxl import load_workbook

from config import EXCEL_FILE


class DashboardScreen:

    FILE = EXCEL_FILE

    def __init__(self, parent):
        self.parent = parent
        self.datetime_label = None

        self.total_label = None
        self.applied_label = None
        self.interview_label = None
        self.selected_label = None
        self.rejected_label = None
        self.followup_label = None

        self.table = None

    def read_excel(self):

        result = {
            "Total": 0,
            "Applied": 0,
            "Interview": 0,
            "Selected": 0,
            "Rejected": 0,
            "Followup": 0,
            "Recent": []
        }

        if not os.path.exists(self.FILE):
            return result

        wb = load_workbook(self.FILE)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        if len(rows) <= 1:
            return result

        header = rows[0]
        data = rows[1:]

        result["Total"] = len(data)

        status_index = header.index("Status")
        company_index = header.index("Company")
        role_index = header.index("Role")
        follow_index = header.index("Follow-up Date")

        counter = Counter()

        today = datetime.now().strftime("%d-%m-%Y")

        for row in data:

            status = str(row[status_index]).strip()

            counter[status] += 1

            if row[follow_index]:
                if str(row[follow_index]).strip() == today:
                    result["Followup"] += 1

        result["Applied"] = counter.get("Applied", 0)
        result["Interview"] = counter.get("Interview", 0)
        result["Selected"] = counter.get("Selected", 0)
        result["Rejected"] = counter.get("Rejected", 0)

        recent = []

        for row in data[-5:]:

            recent.append(
                (
                    row[company_index],
                    row[role_index],
                    row[status_index]
                )
            )

        result["Recent"] = recent[::-1]

        return result

    def card(self, parent, title, value):

        box = ctk.CTkFrame(parent, width=170, height=120)

        box.pack(side="left", padx=10)

        ctk.CTkLabel(
            box,
            text=title,
            font=("Arial",18,"bold")
        ).pack(pady=(18,5))

        lbl = ctk.CTkLabel(
            box,
            text=value,
            font=("Arial",30,"bold")
        )

        lbl.pack()

        return lbl

    def refresh(self):

        data = self.read_excel()

        self.total_label.configure(text=str(data["Total"]))
        self.applied_label.configure(text=str(data["Applied"]))
        self.interview_label.configure(text=str(data["Interview"]))
        self.selected_label.configure(text=str(data["Selected"]))
        self.rejected_label.configure(text=str(data["Rejected"]))
        self.followup_label.configure(text=str(data["Followup"]))

        self.table.configure(state="normal")
        self.table.delete("1.0","end")

        self.table.insert(
            "end",
            f'{"Company":25} {"Role":25} Status\n'
        )

        self.table.insert(
            "end",
            "-"*70+"\n"
        )

        for c,r,s in data["Recent"]:

            self.table.insert(
                "end",
                f"{str(c):25} {str(r):25} {str(s)}\n"
            )

        self.table.configure(state="disabled")

    def show(self):

        frame = ctk.CTkFrame(self.parent)

        frame.pack(fill="both",expand=True,padx=20,pady=20)

        ctk.CTkLabel(
            frame,
            text="JOB TRACKER PRO DASHBOARD",
            font=("Arial",30,"bold")
        ).pack(pady=15)

        self.datetime_label = ctk.CTkLabel(
            frame,
            text="",
            font=("Arial",16)
        )

        self.datetime_label.pack()

        self.update_datetime()

        cards = ctk.CTkFrame(frame)

        cards.pack(pady=25)

        self.total_label = self.card(cards,"Total Jobs","0")
        self.applied_label = self.card(cards,"Applied","0")
        self.interview_label = self.card(cards,"Interview","0")
        self.selected_label = self.card(cards,"Selected","0")
        self.rejected_label = self.card(cards,"Rejected","0")
        self.followup_label = self.card(cards,"Today's Follow-up","0")

        ctk.CTkLabel(
            frame,
            text="Recent Applications",
            font=("Arial",20,"bold")
        ).pack(pady=(20,10))

        self.table = ctk.CTkTextbox(
            frame,
            width=900,
            height=180
        )

        self.table.pack()

        ctk.CTkButton(
            frame,
            text="Refresh Dashboard",
            command=self.refresh,
            width=220,
            height=40
        ).pack(pady=20)

        self.refresh()

        return frame

    def update_datetime(self):

        if self.datetime_label:

            now = datetime.now().strftime("%d-%b-%Y   %I:%M:%S %p")

            self.datetime_label.configure(
                text=f"📅 {now}"
            )

            self.parent.after(
                1000,
                self.update_datetime
            )